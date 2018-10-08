#####################################################################################################################################################################################################
    #3PAM Extract Validator
    #Input : .xlsx file extracted from Chevron 3Pam system
    #Output: Text indicating if columns of input file are correct as required for 432 Data Migration transform script
#####################################################################################################################################################################################################


from openpyxl import Workbook, load_workbook
import datetime
import sys

filename = sys.argv[1]  #Filename is expected to be the first argument on the commandline.

def checkColumn(checkColumn, expectedValue):
    """Given a column number and an expected value return true/false indicating if that value is present at column"""
    returnValue = False
    if ws.cell(row=1, column=checkColumn).value == expectedValue:
        returnValue = True
    return returnValue

#def runTest(column, value):
#    if checkColumn(column,value):
#        print("Column " + str(column) + " is correct.")
#    else:
#        print("Column " + str(column) + " is not correct!  Expected " + value)

def runTest(column, value):
    if checkColumn(column,value):
        print("Column " + numberToLetter(column) + " is correct.")
    else:
        print("Column " + numberToLetter(column) + " is not correct!  Expected " + value)



def numberToLetter(colNumber):
    """Given a column number return the excel column heading"""
    ascii = colNumber + 64
    return chr(ascii)

#load workbook
print("Loading: " + filename)
wb = load_workbook(filename)

#activate sheet
ws = wb.active

runTest(1,"EntityID")
runTest(2,"EntityName")
runTest(3,"EntityCountry")
runTest(4,"EntityCountryName")
runTest(5,"EntityAliasName")
runTest(6,"ParentEntityName")
runTest(7,"SourcingCompanyName")
runTest(8,"SourcingCompanyID")
runTest(9,"EntityTechData")
runTest(10,"EntitySPData")
runTest(11,"EntityIPData")
runTest(12,"ResourceID")
runTest(13,"ResourceType")
runTest(14,"ResourceName")
runTest(15,"ResourceCountry")
runTest(16,"ResourceDescription")
runTest(17,"ResourceURL")
runTest(18,"ResourceECCN")
runTest(19,"ApprovedAccessOPCO")
runTest(20,"ApprovedAccessApprovalDate")
runTest(21,"ApprovedAccessSourceInfo")

