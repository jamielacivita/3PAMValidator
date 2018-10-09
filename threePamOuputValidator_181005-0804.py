#####################################################################################################################################################################################################
    #3PAM Extract Validator
    #Input : .xlsx file extracted from Chevron 3Pam system
    #Output: Text indicating if columns of input file are correct as required for 432 Data Migration transform script
#####################################################################################################################################################################################################


from openpyxl import Workbook, load_workbook
import datetime
import sys

filename = sys.argv[1]  #Filename is expected to be the first argument on the commandline.

def checkColumn(ws, checkColumn, expectedValue):
    """Given a column number and an expected value return true/false indicating if that value is present at column"""
    returnValue = False
    if ws.cell(row=1, column=checkColumn).value == expectedValue:
        returnValue = True
    return returnValue


def runTest(ws, column, value):
    if checkColumn(ws, column,value):
        print("Column " + numberToLetter(column) + " is correct.")
    else:
        print("Column " + numberToLetter(column) + " is not correct!  Expected " + value)



def numberToLetter(colNumber):
    """Given a column number return the excel column heading"""
    ascii = colNumber + 64
    return chr(ascii)


def main():
    #load workbook
    print("Loading: " + filename)
    wb = load_workbook(filename)

    #activate sheet
    ws = wb.active

    headerContents = {}
    headerContents[1] = "EntityID"
    headerContents[2] = "EntityName"
    headerContents[3] = "EntityCountry"
    headerContents[4] = "EntityCountryName"
    headerContents[5] = "EntityAliasName"
    headerContents[6] = "ParentEntityName"
    headerContents[7] = "SourcingCompanyName"
    headerContents[8] = "SourcingCompanyID"
    headerContents[9] = "EntityTechData"
    headerContents[10] = "EntitySPData"
    headerContents[11] = "EntityIPData"
    headerContents[12] = "ResourceID"
    headerContents[13] = "ResourceType"
    headerContents[14] = "ResourceName"
    headerContents[15] = "ResourceCountry"
    headerContents[16] = "ResourceDescription"
    headerContents[17] = "ResourceURL"
    headerContents[18] = "ResourceECCN"
    headerContents[19] = "ApprovedAccessOPCO"
    headerContents[20] = "ApprovedAccessApprovalDate"
    headerContents[21] = "ApprovedAccessSourceInfo"


    for c in range(1,22):
        runTest(ws,c,headerContents[c])


if (__name__ == "__main__"):
    main()



